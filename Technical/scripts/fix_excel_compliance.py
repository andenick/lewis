"""
Excel Compliance Validator and Fixer
====================================

This script validates Excel files for one-sheet compliance and
automatically fixes multi-sheet violations by splitting them into
single-sheet files.

Standard: EVERY Excel output file MUST have EXACTLY one sheet.

Author: Lewis Platform
Date: October 14, 2025
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import List, Dict, Tuple
import openpyxl
import warnings
warnings.filterwarnings('ignore')

class ExcelComplianceFixer:
    """Validate and fix Excel one-sheet compliance violations."""

    def __init__(self, data_directory: Path = None):
        """Initialize the compliance fixer."""
        if data_directory is None:
            # Auto-detect Lewis project data directory
            self.data_dir = Path(__file__).parent.parent.parent / "Output" / "Data"
        else:
            self.data_dir = Path(data_directory)

        self.violations = []
        self.fixes_applied = []

    def validate_all_excel_files(self) -> Dict[str, List[str]]:
        """
        Validate all Excel files in the data directory for one-sheet compliance.

        Returns:
            Dictionary with 'compliant' and 'violations' file lists
        """
        print("🔍 VALIDATING EXCEL ONE-SHEET COMPLIANCE")
        print("=" * 60)

        excel_files = list(self.data_dir.rglob("*.xlsx"))
        compliant_files = []
        violating_files = []

        for file_path in excel_files:
            try:
                xl = pd.ExcelFile(file_path)
                sheet_count = len(xl.sheet_names)

                if sheet_count == 1:
                    compliant_files.append(str(file_path))
                    print(f"✅ COMPLIANT: {file_path.name} (1 sheet)")
                else:
                    violating_files.append(str(file_path))
                    self.violations.append({
                        'file': str(file_path),
                        'sheets': xl.sheet_names,
                        'count': sheet_count
                    })
                    print(f"❌ VIOLATION: {file_path.name} ({sheet_count} sheets: {xl.sheet_names})")

            except Exception as e:
                print(f"⚠️  ERROR reading {file_path.name}: {e}")
                violating_files.append(str(file_path))

        print("\n" + "=" * 60)
        print(f"SUMMARY: {len(compliant_files)} compliant, {len(violating_files)} violations")

        return {
            'compliant': compliant_files,
            'violations': violating_files
        }

    def fix_world_bank_gdp_file(self) -> bool:
        """
        Fix BoP_WBankGDP_NA.xlsx by splitting into 2 single-sheet files.

        Returns:
            True if fix was successful
        """
        print("\n🔧 FIXING: BoP_WBankGDP_NA.xlsx")

        # Find the file
        gdp_file = None
        for violation in self.violations:
            if 'BoP_WBankGDP_NA.xlsx' in violation['file']:
                gdp_file = Path(violation['file'])
                break

        if not gdp_file:
            print("❌ BoP_WBankGDP_NA.xlsx not found in violations")
            return False

        try:
            # Read all sheets
            xl = pd.ExcelFile(gdp_file)
            print(f"Found sheets: {xl.sheet_names}")

            # Process each sheet
            for sheet_name in xl.sheet_names:
                df = pd.read_excel(gdp_file, sheet_name=sheet_name)

                # Create new filename based on sheet content
                if sheet_name.lower() in ['81000-0001', 'data', 'gdp_data']:
                    output_file = gdp_file.parent / "world_bank_gdp_data.xlsx"
                elif sheet_name.lower() in ['note', 'notes', 'metadata']:
                    output_file = gdp_file.parent / "world_bank_gdp_notes.xlsx"
                else:
                    # Generic naming
                    safe_name = sheet_name.replace(' ', '_').replace('-', '_').lower()
                    output_file = gdp_file.parent / f"world_bank_gdp_{safe_name}.xlsx"

                # Write single-sheet Excel file
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Data', index=False)

                # Format the sheet
                self._format_excel_sheet(output_file)

                print(f"✅ Created: {output_file.name}")
                self.fixes_applied.append(f"Split {gdp_file.name} sheet '{sheet_name}' -> {output_file.name}")

            # Remove original multi-sheet file
            gdp_file.unlink()
            print(f"🗑️  Removed original: {gdp_file.name}")
            self.fixes_applied.append(f"Removed original multi-sheet file: {gdp_file.name}")

            return True

        except Exception as e:
            print(f"❌ Error fixing {gdp_file.name}: {e}")
            return False

    def fix_trade_analysis_file(self) -> bool:
        """
        Fix trade_analysis_results.xlsx by splitting into 4 themed files.

        Returns:
            True if fix was successful
        """
        print("\n🔧 FIXING: trade_analysis_results.xlsx")

        # Find the file
        trade_file = None
        for violation in self.violations:
            if 'trade_analysis_results.xlsx' in violation['file']:
                trade_file = Path(violation['file'])
                break

        if not trade_file:
            print("❌ trade_analysis_results.xlsx not found in violations")
            return False

        try:
            # Read all sheets
            xl = pd.ExcelFile(trade_file)
            print(f"Found sheets: {xl.sheet_names}")

            # Sheet to filename mapping
            sheet_mappings = {
                'Data_Summary': 'trade_summary_statistics.xlsx',
                'US_Current_Account': 'us_current_account_analysis.xlsx',
                'US_Period_Stats': 'us_period_statistics.xlsx',
                'Cross_Country_Comparison': 'cross_country_comparison.xlsx'
            }

            # Process each sheet
            for sheet_name in xl.sheet_names:
                df = pd.read_excel(trade_file, sheet_name=sheet_name)

                # Determine output filename
                output_name = sheet_mappings.get(sheet_name, f"trade_analysis_{sheet_name.lower().replace(' ', '_')}.xlsx")
                output_file = trade_file.parent / output_name

                # Write single-sheet Excel file
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Data', index=False)

                # Format the sheet
                self._format_excel_sheet(output_file)

                print(f"✅ Created: {output_file.name}")
                self.fixes_applied.append(f"Split {trade_file.name} sheet '{sheet_name}' -> {output_file.name}")

            # Remove original multi-sheet file
            trade_file.unlink()
            print(f"🗑️  Removed original: {trade_file.name}")
            self.fixes_applied.append(f"Removed original multi-sheet file: {trade_file.name}")

            return True

        except Exception as e:
            print(f"❌ Error fixing {trade_file.name}: {e}")
            return False

    def _format_excel_sheet(self, file_path: Path):
        """Format Excel sheet for professional appearance."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = load_workbook(file_path)
            ws = wb.active

            # Format header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)

        except Exception as e:
            print(f"⚠️  Warning: Could not format {file_path.name}: {e}")

    def run_full_fix(self) -> bool:
        """
        Run complete validation and fix process.

        Returns:
            True if all violations were fixed
        """
        print("🚀 STARTING EXCEL COMPLIANCE FIX PROCESS")
        print("=" * 60)

        # Step 1: Validate all files
        validation_results = self.validate_all_excel_files()

        if not validation_results['violations']:
            print("🎉 ALL FILES COMPLIANT - No fixes needed!")
            return True

        # Step 2: Fix specific violations
        fixes_success = True

        # Fix World Bank GDP file
        if any('BoP_WBankGDP_NA.xlsx' in v for v in validation_results['violations']):
            success = self.fix_world_bank_gdp_file()
            fixes_success = fixes_success and success

        # Fix trade analysis file
        if any('trade_analysis_results.xlsx' in v for v in validation_results['violations']):
            success = self.fix_trade_analysis_file()
            fixes_success = fixes_success and success

        # Step 3: Re-validate
        print("\n🔄 RE-VALIDATING AFTER FIXES...")
        print("=" * 60)

        # Clear violations list and re-validate
        self.violations.clear()
        final_results = self.validate_all_excel_files()

        # Step 4: Summary
        print("\n" + "=" * 60)
        print("📊 FIX SUMMARY")
        print("=" * 60)

        for fix in self.fixes_applied:
            print(f"✅ {fix}")

        if not final_results['violations']:
            print(f"\n🎉 SUCCESS: All {len(validation_results['violations'])} violations fixed!")
            print(f"✅ {len(final_results['compliant']) + len(validation_results['violations'])} files now compliant")
            return True
        else:
            print(f"\n⚠️  PARTIAL SUCCESS: {len(final_results['violations'])} violations remain")
            return False


def main():
    """Main execution function."""
    print("Excel Compliance Validator and Fixer")
    print("Lewis International Economics Platform Enhancement")
    print("=" * 60)

    # Initialize fixer
    fixer = ExcelComplianceFixer()

    # Run full fix process
    success = fixer.run_full_fix()

    if success:
        print("\n🎯 ALL EXCEL FILES NOW COMPLY WITH DRUCK STANDARDS!")
    else:
        print("\n⚠️  Some issues remain - manual intervention may be needed")

    return success


if __name__ == "__main__":
    main()